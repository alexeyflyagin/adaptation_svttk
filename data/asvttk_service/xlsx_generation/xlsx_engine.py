import asyncio
import os
import platform
from concurrent.futures import ThreadPoolExecutor

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typeguard import typechecked

from data.asvttk_service.xlsx_generation.types import ReportTable, ReportFile

HEADER_HEIGHT = 30.0
HEADER_BOLD_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(vertical="bottom")
HEADER_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HEADER_BORDER = Border(left=Side(style="thin", color="808080"), right=Side(style="thin", color="808080"),
                       top=Side(style="thin", color="808080"), bottom=Side(style="thin", color="808080"))


@typechecked
def __create_xlsx(file_name: str, table_types: list[type[ReportTable]], tables: list[ReportTable]) -> str:
    wb = Workbook()
    if table_types:
        wb.remove(wb.active)
    for table_type in table_types:
        ws: Worksheet = wb.create_sheet(title=table_type.__tablename__)
        columns = list(table_type.__columns__.values())
        ws.append([i.alias for i in columns])
        ws.row_dimensions[1].height = HEADER_HEIGHT
        for idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=idx)
            cell.font = HEADER_BOLD_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = column.width
    for table in tables:
        ws: Worksheet = wb[table.__tablename__]
        ws.append(table.__values_converted__)
        row = list(ws.rows)[-1]
        for i in range(len(row)):
            style = table.__named_styles__[i]
            if style:
                row[i].style = style

    filename = os.path.join('data', 'asvttk_service', 'xlsx_generation', 'generated', f'{file_name}.xlsx')
    wb.save(filename)
    return filename


async def __run_sync_code_in_thread(*args):
    loop = asyncio.get_running_loop()
    with ThreadPoolExecutor() as pool:
        result = await loop.run_in_executor(pool, __create_xlsx, *args)
    return result


@typechecked
async def create_xlsx(file_name: str, table_types: list[type[ReportTable]], tables: list[ReportTable]) -> ReportFile:
    filename = await __run_sync_code_in_thread(file_name, table_types, tables)
    path = os.path.dirname(filename)
    return ReportFile(path, os.path.basename(filename))
